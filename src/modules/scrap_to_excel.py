def scrap_to_excel():
    from os import system as cmd
    import pandas as pd
    from openpyxl.reader.excel import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    from openpyxl.styles.borders import Border, Side
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.service import Service as ChromeService
    from webdriver_manager.chrome import ChromeDriverManager

    cmd('cls')

    # Importando a planilha modelo
    Path = input("Insira o caminho da planilha.: ")
    BaseDado = pd.read_excel(rf'{Path}') 

    # Criando colunas vazias no DataFrame
    col_names = ["ValorFornecedor", "Pos 1",
                "Pos 2", "Pos 3", "Pos 4", "Pos 5", "Status"]

    # ANOTAÇÃO: Necessito modificar isso
    for col in col_names:
        BaseDado.loc[0, col] = None

    # Informando a licitação de interesse na coleta
    Pregao = str(input("Digite o número do pregão(Ex:12021): "))

    # Armazenando o link para ser aberto no navegador posteriormente
    url = "http://comprasnet.gov.br/acesso.asp?url=/livre/Pregao/lista_pregao_filtro.asp?Opc=2"

    # Informando o Path para o arquivo "chromedriver.exe" e armazenando em "driver"
    op = webdriver.ChromeOptions()
    op.add_argument("--start-maximized")
    op.add_experimental_option("excludeSwitches", ["enable-logging"])
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=op)

    # Abrindo o google chrome no site informado
    driver.get(url)

    # Navegador em tela cheia
    driver.maximize_window()

    # Configurando o tempo de espera em segundos até que o elemento html seja encontrado
    wdw = WebDriverWait(driver, 2000)
    wdw2 = WebDriverWait(driver, 2000)

    # Mudando o main frame do site que será navegado
    driver.switch_to.frame("main2")


    CodUasg = wdw.until(
        EC.presence_of_element_located((By.XPATH, "//input[@id='co_uasg']"))
    )
    CodUasg.send_keys("150182")

    NumPreg = wdw.until(EC.presence_of_element_located(
        (By.XPATH, "//input[@id='numprp']")))
    NumPreg.send_keys(Pregao)

    driver.execute_script("ValidaForm();")
    sleep(1)

    PrClick = wdw.until(
        EC.presence_of_element_located(
            (By.XPATH, "//a[contains(text(), '{}')]".format(Pregao))
        )
    )
    PrClick.click()

    print("Preencha o CAPTCHA")

    wdw.until(
        EC.presence_of_element_located(
            (By.XPATH, "//center[contains(text(), 'MINISTÉRIO DA EDUCAÇÃO')]")
        )
    )
    print("Portal de compras encontrado")

    fail = ["Item deserto", "Cancelado no julgamento"]

    ############################## Alterei aqui ##############################
    item, collecting, page, n = [1, True, 0, len(BaseDado.iloc[:, 0])]

    while collecting:
        actual_item = (page)*100+item
        print(f'\n\n\n\n \U0001F3F8 Buscando pelo item: {actual_item}\n\n')

        Vencedores = wdw2.until(
            EC.presence_of_element_located(
                (
                    By.XPATH,
                    "//tr[contains(@class, 'tex3a')][{}]/td[6]".format(
                        item)
                ),
            )
        )

        if Vencedores.text in fail:
            print("Status do item ", actual_item, ":", Vencedores.text)
            BaseDado.iloc[actual_item-1, 11] = Vencedores.text
            if actual_item == n:
                collecting = False
            elif item == 100:
                item -= 99
                page += 1
                driver.execute_script("javascript:PaginarItens('Proxima');")
            else:
                item += 1
        else:
            Vencedores = wdw2.until(
                EC.presence_of_element_located(
                    (
                        By.XPATH,
                        "//tr[./td/a[contains(@href, 'javascript:void(0)')]][{}]/td[6]/a".format(
                            item),
                    )
                )
            )

            print("Status do item ", actual_item, ":", Vencedores.text, "\n")
            BaseDado.iloc[actual_item-1, 11] = Vencedores.text
            Vencedores.click()
            driver.switch_to.window(driver.window_handles[1])
            Captcha = wdw2.until(
                EC.presence_of_element_located((By.XPATH, "//h2")))
            if Captcha.text == "ACOMPANHAMENTO DE PREGÃO":
                print("Preencha o CAPTCHA")
                wdw.until(EC.presence_of_element_located(
                    (By.CLASS_NAME, "pregao")))

            vencedor = 1
            for colocado in range(1, 6):
                try:
                    InfoEmpresa = []

                    Cnpj = driver.find_element(By.XPATH,
                        "//tr[@class = 'tex3'][{}]/td[1]".format(colocado)
                    )

                    Nome = driver.find_element(By.XPATH,
                        "//tr[@class = 'tex3'][{}]/td[2]".format(colocado)
                    )

                    ValorEmpresa = driver.find_element(By.XPATH,
                        "//tr[@class = 'tex3'][{}]/td[4]".format(colocado)
                    ).text

                    StatusEmpresa = driver.find_element(By.XPATH,
                        "//tr[@class='tex3'][{}]/td[@class='tex3b']".format(
                            colocado)
                    ).text

                    ProdutoInfo = driver.find_element(By.XPATH,
                        "//tr[@class='tex5a'][{}]/td[@colspan='6']".format(
                            colocado)
                    )
                    ProdutoInfo = ProdutoInfo.text.split()

                    Marca = []
                    Fabricante = []
                    Versão = []
                    i = 0
                    while ProdutoInfo[i] != "Descrição":
                        if ProdutoInfo[i] == "Marca:":
                            PosMarca = i
                        elif ProdutoInfo[i] == "Fabricante:":
                            PosFabricante = i
                        elif ProdutoInfo[i] == "Modelo":
                            PosVersão = i
                        i += 1

                    for j in range(PosMarca + 1, PosFabricante):
                        Marca.append(ProdutoInfo[j])
                    for j in range(PosFabricante + 1, PosVersão):
                        Fabricante.append(ProdutoInfo[j])
                    for j in range(PosVersão + 3, i):
                        Versão.append(ProdutoInfo[j])

                    Marca = " ".join(Marca)
                    Fabricante = " ".join(Fabricante)
                    Versão = " ".join(Versão)

                    InfoEmpresa = [
                        ("CNPJ: " + Cnpj.text),
                        ("NOME: " + Nome.text),
                        ("R$ " + ValorEmpresa),
                        ("Marca: " + Marca),
                        ("Fabricante: " + Fabricante),
                        ("Modelo / Versão: " + Versão),
                    ]

                    BaseDado.iloc[actual_item-1, (5 + colocado)] = (
                        InfoEmpresa[0]
                        + "\n"
                        + InfoEmpresa[1]
                        + "\n"
                        + InfoEmpresa[2]
                        + "\n"
                        + InfoEmpresa[3]
                        + "\n"
                        + InfoEmpresa[4]
                        + "\n"
                        + InfoEmpresa[5]
                    )

                    print(InfoEmpresa)

                    if StatusEmpresa == "Recusado":
                        vencedor += 1
                        BaseDado.iloc[actual_item-1, (5 + colocado)] = "Recusado \U0001F6AB \n" + \
                            BaseDado.iloc[actual_item-1, (5 + colocado)]
                        print(f"{StatusEmpresa} '\U0001F6AB'")
                    elif StatusEmpresa == "Aceito":
                        BaseDado.iloc[actual_item-1, (5 + colocado)] = "Aceito \U00002705 \n" + \
                            BaseDado.iloc[actual_item-1, (5 + colocado)]
                        print(f"{StatusEmpresa} '\u2705'")
                    elif StatusEmpresa == "Adjudicado":
                        BaseDado.iloc[actual_item-1, (5 + colocado)] = "Adjudicado \U00002705 \n" + \
                            BaseDado.iloc[actual_item-1, (5 + colocado)]
                        print(f"{StatusEmpresa} '\u2705'")

                    if colocado == vencedor:
                        try:
                            ValorNegociado = driver.find_element(By.XPATH,
                                "//tr[@class='tex3'][{}]/td[not(node())]".format(colocado))
                            BaseDado.iloc[actual_item-1, 5] = float(
                                ValorEmpresa.replace(".", "").replace(",", "."))
                        except:
                            ValorNegociado = driver.find_element(By.XPATH,
                                "//tr[@class='tex3'][{}]/td[6]".format(colocado)).text
                            BaseDado.iloc[actual_item-1, 5] = float(
                                ValorNegociado.replace(".", "").replace(",", "."))
                            BaseDado.iloc[actual_item-1,
                                        (5 + colocado)] = f"Valor Negociado: {ValorNegociado}\n"+BaseDado.iloc[actual_item-1, (5 + colocado)]

                except:
                    break

            driver.switch_to.window(driver.window_handles[0])
            driver.switch_to.frame("main2")
            if actual_item == n:
                collecting = False
            elif item == 100:
                item -= 99
                page += 1
                driver.execute_script("javascript:PaginarItens('Proxima');")
            else:
                item += 1
            BaseDado.to_excel(Path, index=False)

    driver.quit

    def highlight_price(s):
        color = 'red'
        return 'background-color: %s' % color


    def highlight_desert(s):
        color = 'orange'
        return 'background-color: %s' % color


    def highlight_canceled(s):
        color = '#d65f5f'
        return 'background-color: %s' % color


    def highlight_refused(s):
        color = '#8147d1'
        return 'background-color: %s' % color


    def highlight_acepted(s):
        color = '#5cde35'
        return 'background-color: %s' % color


    def highlight_awarded(s):
        color = '#eb38af'
        return 'background-color: %s' % color


    BaseDado = BaseDado.style.applymap(
        highlight_price, subset=pd.IndexSlice[list(BaseDado.query(
            'ValorReferência < ValorFornecedor').index), 'ValorFornecedor']
    ).applymap(
        highlight_desert, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Status'] == 'Item deserto'].index), 'Status']
    ).applymap(
        highlight_canceled, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Status'] == 'Cancelado no julgamento'].index), 'Status']
    ).applymap(
        highlight_refused, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Pos 1'].astype(str).str.contains("Recusado", na=False)].index), 'Pos 1']
    ).applymap(
        highlight_refused, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Pos 2'].astype(str).str.contains("Recusado", na=False)].index), 'Pos 2']
    ).applymap(
        highlight_refused, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Pos 3'].astype(str).str.contains("Recusado", na=False)].index), 'Pos 3']
    ).applymap(
        highlight_refused, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Pos 4'].astype(str).str.contains("Recusado", na=False)].index), 'Pos 4']
    ).applymap(
        highlight_refused, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Pos 5'].astype(str).str.contains("Recusado", na=False)].index), 'Pos 5']
    ).applymap(
        highlight_acepted, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Pos 1'].astype(str).str.contains("Aceito", na=False)].index), 'Pos 1']
    ).applymap(
        highlight_acepted, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Pos 2'].astype(str).str.contains("Aceito", na=False)].index), 'Pos 2']
    ).applymap(
        highlight_acepted, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Pos 3'].astype(str).str.contains("Aceito", na=False)].index), 'Pos 3']
    ).applymap(
        highlight_acepted, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Pos 4'].astype(str).str.contains("Aceito", na=False)].index), 'Pos 4']
    ).applymap(
        highlight_acepted, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Pos 5'].astype(str).str.contains("Aceito", na=False)].index), 'Pos 5']
    ).applymap(
        highlight_awarded, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Pos 1'].astype(str).str.contains("Adjudicado", na=False)].index), 'Pos 1']
    ).applymap(
        highlight_awarded, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Pos 2'].astype(str).str.contains("Adjudicado", na=False)].index), 'Pos 2']
    ).applymap(
        highlight_awarded, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Pos 3'].astype(str).str.contains("Adjudicado", na=False)].index), 'Pos 3']
    ).applymap(
        highlight_awarded, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Pos 4'].astype(str).str.contains("Adjudicado", na=False)].index), 'Pos 4']
    ).applymap(
        highlight_awarded, subset=pd.IndexSlice[list(
            BaseDado[BaseDado['Pos 5'].astype(str).str.contains("Adjudicado", na=False)].index), 'Pos 5']
    )

    BaseDado.to_excel(Path, index=False)

    print('\nExtração Concluída \U0001F7E2')
    print('\nIniciando a estruturação da planilha... \U000023F3')

    ####### INÍCIO DO OPENPYXL ########

    wb = load_workbook(filename = rf'{Path}')

    ws = wb.active

    ws['C1'] = "UNIDADE"
    ws['D1'] = "QUANTIDADE (PROAES)"
    ws['E1'] = "VALOR DE REFERÊNCIA (unitário) (R$)"
    ws['F1'] = "VALOR FIRMAS"
    ws['G1'] = "POSIÇÃO: 1"
    ws['H1'] = "POSIÇÃO: 2"
    ws['I1'] = "POSIÇÃO: 3"
    ws['J1'] = "POSIÇÃO: 4"
    ws['K1'] = "POSIÇÃO: 5"

    ws.sheet_view.zoomScale = 70

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    # ao inserir, tratar as colunas como números iniciando em 1
    def modify(index, ByRow=False, hOrientation='left', vOrientation='top', wrap_text=False, row_space=None, col_space=None, currency=False, border=False, 
        font_name='Calibri', font_size='10', font_color='000000', font_italic=False, font_bold=False, cell_color=None):
        if type(index)==int:
            index = [index]
        row_count = ws.max_row
        column_count = ws.max_column
        for vector in index:
            if ByRow:
                for col in range(1, column_count+1):
                    currentCell = ws[f'{get_column_letter(col)}{vector}']
                    currentCell.alignment = Alignment(horizontal = hOrientation, vertical = vOrientation, wrapText=wrap_text)
                    if border==True:
                        currentCell.border = thin_border
                    ft = Font(name=font_name, size=font_size, color=font_color, italic=font_italic, bold=font_bold)
                    currentCell.font = ft
                    if cell_color is not None:
                        currentCell.fill = PatternFill("solid", start_color=cell_color)
                    if col_space is not None:
                        letter = get_column_letter(col)
                        ws.column_dimensions[letter].width = col_space
                if row_space is not None:
                    ws.row_dimensions[vector].height = row_space
            else:
                for row in range(2, row_count+1):
                    currentCell = ws[f'{get_column_letter(vector)}{row}']
                    currentCell.alignment = Alignment(horizontal = hOrientation, vertical = vOrientation, wrapText=wrap_text)
                    if currency==True:
                        currentCell.number_format = 'R$ #,##0.00'
                    if border==True:
                        currentCell.border = thin_border
                    ft = Font(name=font_name, size=font_size, color=font_color, italic=font_italic, bold=font_bold)
                    currentCell.font = ft
                    if cell_color is not None:
                        currentCell.fill = PatternFill("solid", start_color=cell_color)
                    if row_space is not None:
                        ws.row_dimensions[row].height = row_space
                if col_space is not None:
                    letter = get_column_letter(vector)
                    ws.column_dimensions[letter].width = col_space
                

    modify(index=1, ByRow=True, hOrientation='center', vOrientation='center', wrap_text=True, font_bold=True, row_space=60, cell_color='f5425a')
    modify(index=1, hOrientation='center', vOrientation='center', cell_color='A5B0A2', font_bold=True, row_space=140, col_space=7, border=True)
    modify(index=2, border=True, wrap_text=True, col_space=34)
    modify(index=[3, 4], border=True, col_space=14, hOrientation='center', vOrientation='center')
    modify(index=[5, 6], border=True, col_space=14, hOrientation='center', vOrientation='center', currency=True)
    modify(index=[7,8,9,10,11], border=True, col_space=30, hOrientation='center', vOrientation='center', wrap_text=True)
    modify(index=12, border=True, col_space=22, hOrientation='center', vOrientation='center')

    wb.save(Path)

    print('\nEstruturação Concluida! \U0001F47E')